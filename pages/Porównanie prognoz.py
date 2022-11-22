
from ast import arg
import pandas as pd
import streamlit as st
import openpyxl as xl
import plotly.express as px
from datetime import datetime
import plotly.graph_objects as go
from traitlets import default

from streamlit_option_menu import option_menu


def get_sheetnames_xlsx(filepath):
    wb = xl.load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def rok(x):
    return x.year


st.set_page_config(layout="wide")
#####################################
#PRZYGOTOWANIE DANYCH
#####################################

#####################
#NBP
#####################
#Ważne żeby zakładki w excelu były ułożone równo tak żeby najnowsza była na początku (piersza po lewej)

NBP_g = ()
x = get_sheetnames_xlsx(r'd:\HUKP\Desktop\Workspace\Raporty\Streamlit\PKB\pages\Porównanie prognoz\NBP.xlsx')
x.reverse()
for i in x:
    tmp_df = pd.read_excel(r'd:\HUKP\Desktop\Workspace\Raporty\Streamlit\PKB\pages\Porównanie prognoz\NBP.xlsx', sheet_name=i)
    tmp_df = tmp_df.rename(columns={"Unnamed: 0": "Zmienna"})
    tmp_df = tmp_df.iloc[:,[0,tmp_df.shape[1]-4,tmp_df.shape[1]-3,tmp_df.shape[1]-2,tmp_df.shape[1]-1]]
    tmp_df.index = tmp_df["Zmienna"]
    tmp_df = tmp_df.drop('Zmienna', axis=1)
    tmp_df = tmp_df.T
    tmp_df = tmp_df.iloc[:,[0,4,5,6,7,8,9,10,11,13,14]]
    tmp_df = tmp_df.set_axis(['CPI','PKB','Popyt_krajowy','PCR','GCR','ITR','wklad_eks_netto','XTR','MTR','LN','UR'], axis=1, inplace=False)
    tmp_df["Prognoza"] = "NBP - {}".format(i)
    tmp_df["Rok"] = tmp_df.index
    #nie jestem pewny czy dobrze liczę udział zapasów w dynamice PKB
    tmp_df["Zapasy"] = tmp_df["PKB"] - (tmp_df["Popyt_krajowy"] + tmp_df["wklad_eks_netto"])
    tmp_df = tmp_df.round(1)

    NBP_g = NBP_g + tuple((tmp_df,))

#####################
#KE
#####################
KE_g = ()
x = get_sheetnames_xlsx(r'd:\HUKP\Desktop\Workspace\Raporty\Streamlit\PKB\pages\Porównanie prognoz\KE.xlsx')
x.reverse()
for i in x:
    tmp_df = pd.read_excel(r'd:\HUKP\Desktop\Workspace\Raporty\Streamlit\PKB\pages\Porównanie prognoz\KE.xlsx', sheet_name=i)
    tmp_df.index = tmp_df["Kategoria"]
    tmp_df = tmp_df.drop('Kategoria',axis=1)
    tmp_df = tmp_df.T
    tmp_df = tmp_df.iloc[:,[0,1,2,3,4,5,7,8,9,10,11,16]]
    tmp_df = tmp_df.set_axis(["PKB","PCR","GCR","ITR","XTR","MTR","Popyt_krajowy","Zapasy","wklad_eks_netto","LN","UR","CPI"], axis=1, inplace=False)
    tmp_df = tmp_df.round(1)
    tmp_df["Rok"] = tmp_df.index
    tmp_df["Prognoza"] = "KE - {}".format(i)
    

    KE_g = KE_g + tuple((tmp_df,))


#####################
#MF
#####################
MF_g = ()
x = get_sheetnames_xlsx(r'd:\HUKP\Desktop\Workspace\Raporty\Streamlit\PKB\pages\Porównanie prognoz\MF.xlsx')
x.reverse()
for i in x:
    tmp_df = pd.read_excel(r'd:\HUKP\Desktop\Workspace\Raporty\Streamlit\PKB\pages\Porównanie prognoz\MF.xlsx', sheet_name=i)
    tmp_df["Prognoza"] = "MF - {}".format(i)
    tmp_df = tmp_df.round(1)
    MF_g = MF_g + tuple((tmp_df,))


#####################################
#ŁĄCZENIE DANYCH
#####################################
KE = pd.concat(KE_g, axis=0, ignore_index=True)
NBP = pd.concat(NBP_g, axis = 0, ignore_index=True)
MF = pd.concat(MF_g, axis=0, ignore_index=True)
df = pd.concat([KE,NBP,MF])



##########################################################
#STREAMLIT
##########################################################

selected = option_menu(
    menu_title = None, 
    options = ["Dynamika", "Wkład we wzrost"],
    menu_icon = "cast",
    default_index = 0,
    orientation = "horizontal"
)

st.title("Porównanie prognoz")


st.markdown("""
    Przedstawienie porównania prognoz: KE (Komisji Europejskiej), NBP (Narodowego Banku Polskiego), MF (Ministerstwa Finansów).
""")

if selected == "Dynamika": 


    st.sidebar.header("")

    #wycięto bezrobocie
    zmienne = ["PKB","CPI","Konsumpcja prywatna","Konumpcja publiczna","Inwestycje","Eksport","Import","Zatrudnienie"]
    wybor_zmiennej = st.sidebar.selectbox("Wybór zmiennej",zmienne)

    wybor_prognozy_NBP = st.sidebar.multiselect("Wybór prognozy NBP",NBP["Prognoza"].unique(),NBP["Prognoza"][len(NBP["Prognoza"])-1])
    wybor_prognozy_KE = st.sidebar.multiselect("Wybór prognozy KE",KE["Prognoza"].unique(),KE["Prognoza"][len(KE["Prognoza"])-1])
    wybor_prognozy_MF = st.sidebar.multiselect("Wybór prognozy MF",MF["Prognoza"].unique(),MF["Prognoza"][len(MF["Prognoza"])-1])

    wybor_prognozy = wybor_prognozy_NBP + wybor_prognozy_KE + wybor_prognozy_MF 


    #####################################
    #WYBÓR DANYCH
    #####################################


    dict = {
        "PKB": "PKB",
        "Konsumpcja prywatna":"PCR",
        "Konumpcja publiczna":"GCR",
        "Inwestycje":"ITR",
        "Eksport": "XTR",
        "Import":"MTR",
        "Zatrudnienie":"LN",
        "Bezrobocie":"UR",
        "CPI":"CPI"
    }
    wybrana_zmienna = dict.get(wybor_zmiennej) 


    dynamiki = df[[wybrana_zmienna,"Prognoza","Rok"]][df.Prognoza.isin(wybor_prognozy)]
    dynamiki["Rok"] = pd.to_datetime(dynamiki["Rok"], format='%Y')

    fig = px.line(dynamiki , x="Rok", y=wybrana_zmienna, color='Prognoza', text=wybrana_zmienna, symbol="Prognoza",width=1000, height=500)
    #fig = px.scatter(temp, x="Rok", y=wybor_zmiennej, color="Prognoza")
    #fig = px.histogram(temp, x="Rok", y=wybor_zmiennej,color='Prognoza', barmode='group')
    fig.update_traces(textposition="bottom right")
    fig.update_xaxes(dtick="M12")

    st.plotly_chart(fig, use_container_width=False)





if selected == "Wkład we wzrost":


    wybor_prognozy_NBP = st.sidebar.multiselect("Wybór prognozy NBP",NBP["Prognoza"].unique(),NBP["Prognoza"][len(NBP["Prognoza"])-1])
    wybor_prognozy_KE = st.sidebar.multiselect("Wybór prognozy KE",KE["Prognoza"].unique(),KE["Prognoza"][len(KE["Prognoza"])-1])
    wybor_prognozy_MF = st.sidebar.multiselect("Wybór prognozy MF",MF["Prognoza"].unique(),MF["Prognoza"][len(MF["Prognoza"])-1])

    wybor_prognozy = wybor_prognozy_NBP + wybor_prognozy_KE + wybor_prognozy_MF 

    
    wklad = df[['PKB','Popyt_krajowy','Zapasy','wklad_eks_netto',"Prognoza","Rok"]][df.Prognoza.isin(wybor_prognozy)]
    wklad["Rok"] = pd.to_datetime(wklad["Rok"], format='%Y')

    wklad["Rok"] = wklad["Rok"].apply(rok)


    fig = go.Figure()



    categories = wklad["Prognoza"].unique()

    for j, i in enumerate(categories):
            kk = wklad[wklad["Prognoza"] == i]
            if j == 0:

                    fig.add_trace(
                            go.Bar(x=[kk["Rok"], [i]*len(kk["Rok"])], y=kk["Popyt_krajowy"],marker_color='#2e59a8', name = "Popyt krajowy",text =kk["Popyt_krajowy"]),
                    )

                    fig.add_trace(
                            go.Bar(x=[kk["Rok"], [i]*len(kk["Rok"])], y=kk["Zapasy"],marker_color='#73a2c6', name = "Zapasy",text =kk["Zapasy"]),
                    )

                    fig.add_trace(
                            go.Bar(x=[kk["Rok"], [i]*len(kk["Rok"])], y=kk["wklad_eks_netto"],marker_color='#c5eddf', name = "Eksport netto",text =kk["wklad_eks_netto"]),
                    )
                    fig.add_trace(go.Scatter(x=[kk["Rok"], [i]*len(kk["Rok"])], y=kk["PKB"],
                        mode='markers',
                        name='PKB',marker_color="Red"))

            else:
                    fig.add_trace(
                            go.Bar(x=[kk["Rok"], [i]*len(kk["Rok"])], y=kk["Popyt_krajowy"],marker_color='#2e59a8', name = "Popyt krajowy", showlegend=False,text =kk["Popyt_krajowy"]),
                    )

                    fig.add_trace(
                            go.Bar(x=[kk["Rok"], [i]*len(kk["Rok"])], y=kk["Zapasy"],marker_color='#73a2c6', name = "Zapasy", showlegend=False,text =kk["Zapasy"]),
                    )

                    fig.add_trace(
                            go.Bar(x=[kk["Rok"], [i]*len(kk["Rok"])], y=kk["wklad_eks_netto"],marker_color='#c5eddf', name = "Eksport netto", showlegend=False,text =kk["wklad_eks_netto"]),
                    )
                    fig.add_trace(go.Scatter(x=[kk["Rok"], [i]*len(kk["Rok"])], y=kk["PKB"],
                        mode='markers',
                        name='PKB', showlegend=False,marker_color="Red"))

                    

    fig.update_layout(
        xaxis=dict(title_text="Rok"),
        yaxis=dict(title_text="mln PLN"),
        barmode='relative',
    )

    fig.update_layout(
        xaxis_title="Rok",
        yaxis_title="%",
        legend_title="",
        font=dict(
            family="times new roman",
            size=12,
            color="Black"
        ),height=600, width=300*len(categories)
    )
        

    st.plotly_chart(fig, use_container_width=False)















